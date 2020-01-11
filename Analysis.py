import math
import matplotlib
import matplotlib.pyplot as plt
import openpyxl
from sklearn.linear_model import LinearRegression


wb = openpyxl.load_workbook('Responses.xlsx')
ws = wb.active

list_screen_time = []
list_score = []

for values in range(128):
    cell_value = values + 2
    screen_time = ws[f'G{cell_value}'].value
    list_screen_time.append(screen_time)

for values in range(128):
    cell_value = values + 2
    screen_time = ws[f'U{cell_value}'].value
    list_score.append(screen_time)

print(list_score)
print(list_screen_time)

plt.scatter(list_screen_time, list_score)

plt.title("Screen Time vs. Mental Health Score")
plt.xlabel('Screen Time')
plt.ylabel('Mental Health Score')
plt.show()





