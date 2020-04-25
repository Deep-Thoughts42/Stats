import openpyxl
import globals as g
import os
import time

# I’ve been feeling useful
# I’ve been feeling relaxed
# I’ve had energy to spare
# I’ve been dealing with problems well
# I’ve been thinking clearly
# I’ve been feeling good about myself
# I’ve been feeling close to other people
# I’ve been feeling confident
# I’ve been able to make up my own mind about things
# I’ve been feeling loved
# I’ve been interested in new things
# I’ve been feeling cheerful


print(os.getcwd())

wb = openpyxl.load_workbook('StatsProject.xlsx')
ws = wb.active


questions_list = (
    "I’ve been feeling useful",
    "I’ve been feeling relaxed",
    "I’ve had energy to spare",
    "I’ve been dealing with problems well",
    "I’ve been thinking clearly",
    "I’ve been feeling good about myself",
    "I’ve been feeling close to other people",
    "I’ve been feeling confident",
    "I’ve been able to make up my own mind about things",
    "I’ve been feeling loved",
    "I’ve been interested in new things",
    "I’ve been feeling cheerful"
)

course = {
    "Collings": 1,
    "John": 2,
    "Smith": 3,
    "Walcott": 4,
    "Armstrong": 5,
    "Lee": 6,
    "Warner": 7,
    "Sidhu": 8,
    "Grade 11 Spare": 9,
    "Mehta": 10,
    "Quattrociocchi": 11,
    "Vidulin": 12,
    "Grade 12 Spare": 13,
    "Perry": 14,
    "Sahota": 15,
    "Matei": 16,
}

array = [
    ['Collings', 'Grade 9', 'Open'],
    ['John', 'Grade 9', 'Applied'],
    ['Smith', 'Grade 9', 'Academic'],
    ['Walcott', 'Grade 9', 'AP'],
    ['Armstrong', 'Grade 10', 'Open'],
    ['Lee', 'Grade 10', 'Applied'],
    ['Warner', 'Grade 10', 'Academic'],
    ['Sidhu', 'Grade 10', 'AP'],
    ['Spare', 'Grade 11', 'N/A'],
    ['Mehta', 'Grade 11', 'College'],
    ['Quattrociocchi', 'Grade 11', 'University'],
    ['Vidulin', 'Grade 11', 'AP'],
    ['Spare', 'Grade 12', 'N/A'],
    ['Perry', 'Grade 12', 'College'],
    ['Sahota', 'Grade 12', 'University'],
    ['Matei', 'Grade 12', 'AP']
]

a = ws['a1']

print(a.value)


def dev_setup():
    g.participants = int(input('How many people are taking the survey right now?'))


def start_process():
    print('Thanks for taking part in our stats survey, we will be asking you some questions about your screen time \n'
          'and mental health.')

    for participant in range(g.participants):
        g.gender = input('First off, what is your gender? Type male or female.')

        print("Your teacher has been assigned a number, please type in the number associated with them.")
        time.sleep(3)
        for section in course:
            print(f"{section}: {course[section]}")
            time.sleep(0.2)
        g.teacher_number = int(input("What is your teacher's number?"))
        g.screen_time = int(input('Thank you! Now please type in the amount of time you spend on your phone per week.'))
        mental_health_test()
        print('Thank you very much, that is all we need!')

        ws.insert_rows(2)
        ws['A2'] = array[g.teacher_number-1][0]
        ws['B2'] = array[g.teacher_number-1][1]
        ws['C2'] = array[g.teacher_number - 1][2]
        ws['D2'] = g.gender
        ws['E2'] = g.mental_health_score
        ws['F2'] = g.screen_time

        input("Press any key to finish survey")




def mental_health_test():
    g.mental_health_score = 0
    time.sleep(2)
    print("Now, we will be showing you a series of statements to determine your mental health score. \n"
          "You will have to respond with a 1, 2, 3, 4, or 5 based on how strongly you agree with them.")
    print(' . \n .')
    for question in questions_list:
        print(question)
        print(' . \n . \n .')
        add_value = int(input("From a scale of 1-5, how strongly does this describe you?"))

        g.mental_health_score += add_value
        time.sleep(1)
    print(f'Based upon this experiment, your total mental health score is {g.mental_health_score} out of 60.')


dev_setup()
start_process()
wb.save('StatsProject.xlsx')
